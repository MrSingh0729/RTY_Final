from flask import Blueprint, render_template, request, send_file, jsonify, redirect, url_for
from flask_login import login_required, current_user
import pandas as pd
from models import ModelDescription, ProjectGoal, FPYAutoData
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from utils.api import *
from utils.helpers import get_top_n_counts, map_api_to_db, map_db_to_api
from config import Config
from models import ModelDescription, ProjectGoal, FPYData
from extensions import db

dashboard_bp = Blueprint('dashboard', __name__, url_prefix='/dashboard')

def get_ntf_details_for_station(model_name, station, station_type, start_date, end_date):
    """Get NTF details for a specific station"""
    try:
        token = get_token()
        detail_data = get_station_ntf_details_by_model(token, model_name, station, station_type, start_date, end_date)
        if not detail_data:
            return {
                "top_computers": {},
                "top_faults_by_computer": {}
            }
            
        detail_df = pd.DataFrame(detail_data).rename(columns={
            "substation": "Computer Name",
            "sn": "SN",
            "symptomEnName": "Fault Description"
        })
        detail_df = detail_df[["SN", "Fault Description", "Computer Name"]]
        
        if not detail_df.empty:
            top_computers = detail_df["Computer Name"].value_counts().head(3).to_dict()
            top_faults_by_computer = {}
            
            for comp in top_computers:
                comp_faults = detail_df[detail_df["Computer Name"] == comp]
                faults = comp_faults["Fault Description"].value_counts().head(3).reset_index().values.tolist()
                top_faults_by_computer[comp] = faults
                
            return {
                "top_computers": top_computers,
                "top_faults_by_computer": top_faults_by_computer
            }
        else:
            return {
                "top_computers": {},
                "top_faults_by_computer": {}
            }
    except Exception as e:
        print(f"Error getting NTF details for {station}: {e}")
        return {
            "top_computers": {},
            "top_faults_by_computer": {}
        }

def get_der_details_for_station(model_name, station, station_type, start_date, end_date):
    """Get DER details for a specific station"""
    try:
        token = get_token()
        detail_df = pd.DataFrame(get_station_der_details_by_model(token, model_name, station, station_type, start_date, end_date)).rename(columns={
            "sn": "SN",
            "responsibilityEnName": "Responsibility",
            "symptomEnName": "Symptoms"
        })
        detail_df = detail_df[["SN", "Responsibility", "Symptoms"]]
        
        # Get top symptoms and responsibilities
        if not detail_df.empty:
            top_symptoms = get_top_n_counts(detail_df, "Symptoms", 3)
            top_responsibilities = get_top_n_counts(detail_df, "Responsibility", 3)
            
            return {
                "top_symptoms": top_symptoms.to_dict(orient="records"),
                "top_responsibilities": top_responsibilities.to_dict(orient="records")
            }
        else:
            return {
                "top_symptoms": [],
                "top_responsibilities": []
            }
    except Exception as e:
        print(f"Error getting DER details for {station}: {e}")
        return {
            "top_symptoms": [],
            "top_responsibilities": []
        }

@dashboard_bp.route('/')
def index():
    current_time = datetime.now().strftime('%H:%M')
    try:
        token = get_token()
        projects = get_project_list(token)
        
        # Get today's data from 08:00AM to now
        now = datetime.now()
        start = now.replace(hour=8, minute=0, second=0, microsecond=0)
        
        # Get FPY data for all projects
        fpy_data_raw = get_fpy(token, projects, start, now)
        
        # Get all models from database to filter out 2G models
        db_models = {model.model_name: model.technology for model in ModelDescription.query.all()}
        
        # Initialize counters
        total_pcurr_input = 0
        total_rqc2_good = 0
        valid_rty_count = 0
        rty_sum = 0
        
        # Process each record
        for record in fpy_data_raw:
            project = record.get('project', '')
            station = record.get('station', '')
            technology = db_models.get(project, '')
            
            # Skip 2G models
            if technology == '2G':
                continue
                
            # Extract values
            input_qty = int(record.get('inPut', 0))
            pass_qty = int(record.get('pass', 0))
            rty_val = float(str(record.get('rty', '0')).replace('%', '')) if record.get('rty') else 0
            
            # Accumulate PCURR input
            if station == 'PCURR':
                total_pcurr_input += input_qty
                
            # Accumulate RQC2 good qty
            if station == 'RQC2':
                total_rqc2_good += pass_qty
                
            # Accumulate RTY for average calculation
            if rty_val > 0:
                valid_rty_count += 1
                rty_sum += rty_val
        
        # Calculate average RTY
        avg_rty = (rty_sum / valid_rty_count) if valid_rty_count > 0 else 0
        
        # Get active projects count (excluding 2G)
        active_projects = len([p for p in projects if db_models.get(p, '') != '2G'])
        
        # Prepare stats for template
        stats = {
            'total_pcurr_input': total_pcurr_input,
            'total_rqc2_good': total_rqc2_good,
            'overall_rty': f"{avg_rty:.2f}%",
            'active_projects': active_projects
        }
        
        return render_template('index.html', stats=stats, current_time=current_time)
    except Exception as e:
        current_time = datetime.now().strftime('%H:%M')
        return render_template('errors/500.html', error=str(e), current_time=current_time)
    

@dashboard_bp.route('/home')
def home():
    current_time = datetime.now().strftime('%H:%M')
    # Check if user is logged in
    logged_in = current_user.is_authenticated
    return render_template('home.html', current_time=current_time, logged_in=logged_in)


@dashboard_bp.route('/custom-dashboard')
@login_required
def custom_dashboard():
    current_time = datetime.now().strftime('%H:%M')
    try:
        token = get_token()
        projects = get_project_list(token)
        
        # Get today's data from 08:00AM to now
        now = datetime.now()
        start = now.replace(hour=8, minute=0, second=0, microsecond=0)
        
        # Get FPY data for all projects
        fpy_data_raw = get_fpy(token, projects, start, now)
        
        # Get all models from database to filter out 2G models
        db_models = {model.model_name: model.technology for model in ModelDescription.query.all()}
        
        # Initialize counters
        total_pcurr_input = 0
        total_rqc2_good = 0
        valid_rty_count = 0
        rty_sum = 0
        
        # Process each record
        for record in fpy_data_raw:
            project = record.get('project', '')
            station = record.get('station', '')
            technology = db_models.get(project, '')
            
            # Skip 2G models
            if technology == '2G':
                continue
                
            # Extract values
            input_qty = int(record.get('inPut', 0))
            pass_qty = int(record.get('pass', 0))
            rty_val = float(str(record.get('rty', '0')).replace('%', '')) if record.get('rty') else 0
            
            # Accumulate PCURR input
            if station == 'PCURR':
                total_pcurr_input += input_qty
                
            # Accumulate RQC2 good qty
            if station == 'RQC2':
                total_rqc2_good += pass_qty
                
            # Accumulate RTY for average calculation
            if rty_val > 0:
                valid_rty_count += 1
                rty_sum += rty_val
        
        # Calculate average RTY
        avg_rty = (rty_sum / valid_rty_count) if valid_rty_count > 0 else 0
        
        # Get active projects count (excluding 2G)
        active_projects = len([p for p in projects if db_models.get(p, '') != '2G'])
        
        # Prepare stats for template
        stats = {
            'total_pcurr_input': total_pcurr_input,
            'total_rqc2_good': total_rqc2_good,
            'overall_rty': f"{avg_rty:.2f}%",
            'active_projects': active_projects
        }
        
        return render_template("dashboard/custom_dashboard.html", stats=stats, current_time=current_time)
    except Exception as e:
        current_time = datetime.now().strftime('%H:%M')
        return render_template('errors/500.html', error=str(e), current_time=current_time)

@dashboard_bp.route('/auto-data')
@login_required
def auto_data():
    current_time = datetime.now().strftime('%H:%M')
    try:
        # Try to get data from the auto data table first
        auto_data_records = FPYAutoData.query.all()
        
        if auto_data_records:
            # Convert database objects to dictionaries
            desired_columns = ["project", "station", "inPut", "pass", "fail", "notFail", "der", "ntf", "rty", "py"]
            data = []
            
            for record in auto_data_records:
                row = {
                    "project": record.project,
                    "station": record.station,
                    "inPut": record.inPut,
                    "pass": record.pass_qty,  # Map 'pass_qty' from DB to 'pass' in API
                    "fail": record.fail,
                    "notFail": record.notFail,
                    "der": record.der,
                    "ntf": record.ntf,
                    "rty": record.rty,
                    "py": record.py
                }
                data.append(row)
            
            # Get the last update time
            last_updated = FPYAutoData.query.order_by(FPYAutoData.last_updated.desc()).first().last_updated
            last_updated_str = last_updated.strftime('%Y-%m-%d %H:%M:%S')
        else:
            # If no data in auto data table, fetch fresh data
            print("No auto data found, fetching fresh data...")
            token = get_token()
            projects = get_project_list(token)
            fpy_data_raw = get_fpy(token, projects)
            
            # Process data and add PY column
            desired_columns = ["project", "station", "inPut", "pass", "fail", "notFail", "der", "ntf", "rty", "py"]
            data = []
            
            for row in fpy_data_raw:
                # Add PY column (empty for now)
                row["py"] = ""
                data.append({col: row.get(col, "") for col in desired_columns})
            
            last_updated_str = "Just now"
        
        return render_template('dashboard/auto_data.html', 
                               data=data, 
                               current_time=current_time,
                               last_updated=last_updated_str)
    except Exception as e:
        current_time = datetime.now().strftime('%H:%M')
        return render_template('errors/500.html', error=str(e), current_time=current_time)
    


@dashboard_bp.route('/project-specific', methods=['GET', 'POST'])
@login_required
def project_specific():
    current_time = datetime.now().strftime('%H:%M')
    try:
        token = get_token()
        # Get live running projects from API (as before)
        projects = get_project_list(token)
        selected_project = None
        auto_goal = None
        rty_goal = 90.0  # Default value
        fpy_data = []
        failed_stations = []
        fail_details = []

        if request.method == 'POST':
            selected_project = request.form.get('project')
            # Get goal from database for the selected project
            project_goal = ProjectGoal.query.filter_by(project_name=selected_project).first()
            
            if project_goal and project_goal.goal != 'NA':
                auto_goal = float(project_goal.goal.replace('%', ''))
                rty_goal = auto_goal  # Use the auto-retrieved goal
            else:
                # If no goal in database, use the form value
                rty_goal = float(request.form.get('rty_goal', 90.0))

            print(f"Using RTY goal: {rty_goal}% (Auto: {auto_goal if auto_goal else 'None'})")
            
            fpy_data_raw = get_fpy(token, [selected_project])

            desired_columns = ["project", "station", "inPut", "pass", "fail", "notFail", "der", "ntf", "rty", "py"]
            fpy_data = [
                {col: row.get(col, "") for col in desired_columns}
                for row in fpy_data_raw
            ]

            if fpy_data and "rty" in fpy_data[0]:
                try:
                    actual_rty = float(str(fpy_data[0]["rty"]).replace("%", ""))
                    if actual_rty < rty_goal:
                        for row in fpy_data:
                            station = row.get("station")
                            ntf = float(str(row.get("ntf", "0")).replace("%", "")) if row.get("ntf") else None
                            der = float(str(row.get("der", "0")).replace("%", "")) if row.get("der") else None

                            if station in Config.NTF_GOALS and ntf is not None and ntf > Config.NTF_GOALS[station]:
                                failed_stations.append((station, "NTF", ntf, Config.NTF_GOALS[station]))
                                detail_data = get_station_ntf_details(token, selected_project, station)
                                detail_df = pd.DataFrame(detail_data)
                                detail_df = detail_df.rename(columns={
                                    "substation": "Computer Name",
                                    "sn": "SN",
                                    "symptomEnName": "Fault Description"
                                })
                                detail_df = detail_df[["SN", "Fault Description", "Computer Name"]]

                                top_computers = detail_df["Computer Name"].value_counts().head(3).to_dict()
                                top_faults_by_computer = {}
                                for comp in top_computers:
                                    comp_faults = detail_df[detail_df["Computer Name"] == comp]
                                    faults = comp_faults["Fault Description"].value_counts().head(3).reset_index().values.tolist()
                                    top_faults_by_computer[comp] = faults

                                fail_details.append({
                                    "station": station,
                                    "metric": "NTF",
                                    "actual": ntf,
                                    "goal": Config.NTF_GOALS[station],
                                    "top_computers": top_computers,
                                    "top_faults_by_computer": top_faults_by_computer
                                })

                            if station in Config.DER_GOALS and der is not None and der > Config.DER_GOALS[station]:
                                failed_stations.append((station, "DER", der, Config.DER_GOALS[station]))
                                detail_data = get_station_der_details(token, selected_project, station)
                                detail_df = pd.DataFrame(detail_data)
                                detail_df = detail_df.rename(columns={
                                    "sn": "SN",
                                    "responsibilityEnName": "Responsibility",
                                    "symptomEnName": "Symptoms"
                                })
                                detail_df = detail_df[["SN", "Responsibility", "Symptoms"]]
                                top_symptoms = get_top_n_counts(detail_df, "Symptoms", 3)
                                top_responsibilities = get_top_n_counts(detail_df, "Responsibility", 3)

                                fail_details.append({
                                    "station": station,
                                    "metric": "DER",
                                    "actual": der,
                                    "goal": Config.DER_GOALS[station],
                                    "top_symptoms": top_symptoms.to_dict(orient="records"),
                                    "top_responsibilities": top_responsibilities.to_dict(orient="records")
                                })
                except Exception as e:
                    print("RTY analysis error:", e)

        return render_template("dashboard/project_specific.html",
                               projects=projects,
                               selected_project=selected_project,
                               rty_goal=rty_goal,
                               auto_goal=auto_goal,  # Pass auto goal to template
                               data=fpy_data,
                               failed_stations=failed_stations,
                               fail_details=fail_details,
                               current_time=current_time)
    except Exception as e:
        return render_template('errors/500.html', error=str(e), current_time=current_time)

@dashboard_bp.route('/model-specific', methods=['GET', 'POST'])
@login_required
def model_specific():
    current_time = datetime.now().strftime('%H:%M')
    try:
        # Get all models from database for the searchable dropdown
        models = ModelDescription.query.all()
        selected_model = None
        station_type = "BE"
        start_date = None
        end_date = None
        rty_goal = 90.0  # Default value
        auto_goal = None
        fpy_data = []
        failed_stations = []

        if request.method == 'POST':
            selected_model = request.form.get('model_name')
            station_type = request.form.get('station_type', 'BE')
            start_date = request.form.get('start_date')
            end_date = request.form.get('end_date')
            
            # Get goal from database for the selected model
            model_goal = ModelDescription.query.filter_by(model_name=selected_model).first()
            
            if model_goal and model_goal.goal != 'NA':
                auto_goal = float(model_goal.goal.replace('%', ''))
                rty_goal = auto_goal  # Use the auto-retrieved goal
            else:
                # If no goal in database, use the form value
                rty_goal = float(request.form.get('rty_goal', 90.0))

            print(f"Fetching data for model: {selected_model}, station type: {station_type}, start: {start_date}, end: {end_date}")
            print(f"Using RTY goal: {rty_goal}% (Auto: {auto_goal if auto_goal else 'None'})")
            
            # First, try to get data from the database
            fpy_data_query = FPYData.query.filter_by(project=selected_model)
            
            # If date range is specified, filter by date
            if start_date and end_date:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%dT%H:%M')
                end_datetime = datetime.strptime(end_date, '%Y-%m-%dT%H:%M')
                fpy_data_query = fpy_data_query.filter(FPYData.timestamp.between(start_datetime, end_datetime))
            
            db_data = fpy_data_query.all()
            
            # If no data in database or if refresh is needed, fetch from API
            if not db_data or request.form.get('refresh_data') == 'true':
                token = get_token()
                fpy_data_raw = get_fpy_by_model(token, selected_model, station_type, start_date, end_date)
                print(f"Raw data received: {len(fpy_data_raw) if fpy_data_raw else 0} records")
                
                # Save API data to database
                for record in fpy_data_raw:
                    # Check if record already exists for this model, station, and timestamp
                    existing_record = FPYData.query.filter_by(
                        project=record.get('project'),
                        station=record.get('station')
                    ).first()
                    
                    if existing_record:
                        # Update existing record
                        mapped_data = map_api_to_db(record)
                        for key, value in mapped_data.items():
                            setattr(existing_record, key, value)
                    else:
                        # Create new record
                        mapped_data = map_api_to_db(record)
                        new_record = FPYData(**mapped_data)
                        db.session.add(new_record)
                
                db.session.commit()
                print("Data saved to database")
                
                # Now get the data from the database
                fpy_data_query = FPYData.query.filter_by(project=selected_model)
                if start_date and end_date:
                    start_datetime = datetime.strptime(start_date, '%Y-%m-%dT%H:%M')
                    end_datetime = datetime.strptime(end_date, '%Y-%m-%dT%H:%M')
                    fpy_data_query = fpy_data_query.filter(FPYData.timestamp.between(start_datetime, end_datetime))
                
                db_data = fpy_data_query.all()
            
            # Convert database objects to dictionaries for template
            desired_columns = ["project", "station", "inPut", "pass", "fail", "notFail", "der", "ntf", "rty", "py"]
            fpy_data = []
            
            for record in db_data:
                row = map_db_to_api(record)
                fpy_data.append(row)
            
            print(f"Processed data: {len(fpy_data)} records")
            if fpy_data:
                print(f"First record: {fpy_data[0]}")

            if fpy_data and "rty" in fpy_data[0]:
                try:
                    actual_rty = float(str(fpy_data[0]["rty"]).replace("%", ""))
                    print(f"Actual RTY: {actual_rty}%, Goal: {rty_goal}%")
                    
                    if actual_rty < rty_goal:
                        for row in fpy_data:
                            station = row.get("station")
                            ntf = float(str(row.get("ntf", "0")).replace("%", "")) if row.get("ntf") else None
                            der = float(str(row.get("der", "0")).replace("%", "")) if row.get("der") else None

                            if station in Config.NTF_GOALS and ntf is not None and ntf > Config.NTF_GOALS[station]:
                                failed_stations.append((station, "NTF", ntf, Config.NTF_GOALS[station]))

                            if station in Config.DER_GOALS and der is not None and der > Config.DER_GOALS[station]:
                                failed_stations.append((station, "DER", der, Config.DER_GOALS[station]))
                except Exception as e:
                    print("RTY analysis error:", e)

        return render_template("dashboard/model_specific.html",
                               models=models,
                               selected_model=selected_model,
                               station_type=station_type,
                               start_date=start_date,
                               end_date=end_date,
                               rty_goal=rty_goal,
                               auto_goal=auto_goal,  # Pass auto goal to template
                               data=fpy_data,
                               failed_stations=failed_stations,
                               current_time=current_time,
                               get_ntf_details_for_station=get_ntf_details_for_station,
                               get_der_details_for_station=get_der_details_for_station)
    except Exception as e:
        print(f"Error in model_specific: {e}")
        import traceback
        traceback.print_exc()
        return render_template('errors/500.html', error=str(e), current_time=current_time)

@dashboard_bp.route('/export-excel')
@login_required
def export_excel():
    project = request.args.get('project')
    rty_goal = float(request.args.get('rty_goal', 90.0))
 
    token = get_token()
    fpy_data_raw = get_fpy(token, [project])
 
    if not fpy_data_raw:
        return "No data to export."
 
    # Clean FPY table
    desired_columns = ["project", "station", "inPut", "pass", "fail", "notFail", "der", "ntf", "rty", "py"]
    fpy_data = [{col: row.get(col, "") for col in desired_columns} for row in fpy_data_raw]
    fpy_df = pd.DataFrame(fpy_data).astype(str)
 
    failed_stations = []
    ntf_rows = []
    der_rows = []
 
    try:
        actual_rty = float(str(fpy_df["rty"].iloc[0]).replace("%", ""))
        if actual_rty < rty_goal:
            for _, row in fpy_df.iterrows():
                station = row.get("station")
                ntf = float(str(row.get("ntf", "0")).replace("%", "")) if row.get("ntf") else None
                der = float(str(row.get("der", "0")).replace("%", "")) if row.get("der") else None
 
                if station in Config.NTF_GOALS and ntf is not None and ntf > Config.NTF_GOALS[station]:
                    failed_stations.append((station, "NTF", ntf, Config.NTF_GOALS[station]))
                    detail_df = pd.DataFrame(get_station_ntf_details(token, project, station)).rename(columns={
                        "substation": "Computer Name",
                        "sn": "SN",
                        "symptomEnName": "Fault Description"
                    })[["SN", "Fault Description", "Computer Name"]]
 
                    for comp in detail_df["Computer Name"].value_counts().head(3).index:
                        ntf_rows.append([f"{comp} → {detail_df[detail_df['Computer Name'] == comp]['Computer Name'].count()}", ""])
 
                if station in Config.DER_GOALS and der is not None and der > Config.DER_GOALS[station]:
                    failed_stations.append((station, "DER", der, Config.DER_GOALS[station]))
                    detail_df = pd.DataFrame(get_station_der_details(token, project, station)).rename(columns={
                        "sn": "SN",
                        "responsibilityEnName": "Responsibility",
                        "symptomEnName": "Symptoms"
                    })[["SN", "Responsibility", "Symptoms"]]
 
                    top_symptoms = detail_df["Symptoms"].value_counts().head(3)
                    top_responsibilities = detail_df["Responsibility"].value_counts().head(3)
 
                    for i in range(3):
                        symptom = top_symptoms.index[i] if i < len(top_symptoms) else ""
                        symptom_qty = top_symptoms.iloc[i] if i < len(top_symptoms) else ""
                        resp = top_responsibilities.index[i] if i < len(top_responsibilities) else ""
                        resp_qty = top_responsibilities.iloc[i] if i < len(top_responsibilities) else ""
                        der_rows.append([f"{symptom} → {symptom_qty}", f"{resp} → {resp_qty}"])
    except Exception as e:
        print("RTY analysis error:", e)
 
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "FPY Report"
 
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
 
    def write_section(title, df_or_rows, headers=None):
        # Safely merge title row
        pre_row = ws.max_row if ws.max_row else 0
        ws.append([title])
        title_row = pre_row + 1
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=10)
        ws.cell(row=title_row, column=1).font = bold
 
        if isinstance(df_or_rows, pd.DataFrame):
            ws.append(list(df_or_rows.columns))
            for cell in ws[ws.max_row]:
                cell.font = header_font
                cell.alignment = center
                cell.fill = header_fill
            for row in df_or_rows.itertuples(index=False):
                ws.append(list(row))
        else:
            if headers:
                ws.append(headers)
                for cell in ws[ws.max_row]:
                    cell.font = header_font
                    cell.alignment = center
                    cell.fill = header_fill
            for row in df_or_rows:
                ws.append(row)
 
        ws.append([])  # Spacer row
 
    # Write FPY Table
    write_section("FPY Table", fpy_df)
 
    # Write Failed Stations
    if failed_stations:
        fail_df = pd.DataFrame(failed_stations, columns=["Station", "Metric", "Actual (%)", "Goal (%)"])
        write_section("Failed Stations", fail_df)
 
    # Write NTF Breakdown
    if ntf_rows:
        write_section("Top Failure Analysis — NTF", ntf_rows, ["Top Computer → Qty", "Top 3 Faults → Qty"])
 
    # Write DER Breakdown
    if der_rows:
        write_section("Top Failure Analysis — DER", der_rows, ["Symptom → Qty", "Responsibility → Qty"])
 
    # Adjust column width for 'project' column
    ws.column_dimensions['A'].width = 25  # Wider than others
 
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
 
    return send_file(output,
                     download_name=f"{project}_full_report.xlsx",
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
 
@dashboard_bp.route('/export-pdf')
@login_required
def export_pdf():
    project = request.args.get('project')
    rty_goal = float(request.args.get('rty_goal', 90.0))
 
    token = get_token()
    fpy_data_raw = get_fpy(token, [project])
 
    if not fpy_data_raw:
        return "No data to export."
 
    # Clean FPY table
    desired_columns = ["project", "station", "inPut", "pass", "fail", "notFail", "der", "ntf", "rty", "py"]
    fpy_data = [{col: row.get(col, "") for col in desired_columns} for row in fpy_data_raw]
    fpy_df = pd.DataFrame(fpy_data).astype(str)
 
    failed_stations = []
    ntf_blocks = []
    der_blocks = []
 
    try:
        actual_rty = float(str(fpy_df["rty"].iloc[0]).replace("%", ""))
        if actual_rty < rty_goal:
            for _, row in fpy_df.iterrows():
                station = row.get("station")
                ntf = float(str(row.get("ntf", "0")).replace("%", "")) if row.get("ntf") else None
                der = float(str(row.get("der", "0")).replace("%", "")) if row.get("der") else None
 
                if station in Config.NTF_GOALS and ntf is not None and ntf > Config.NTF_GOALS[station]:
                    failed_stations.append((station, "NTF", ntf, Config.NTF_GOALS[station]))
                    detail_df = pd.DataFrame(get_station_ntf_details(token, project, station)).rename(columns={
                        "substation": "Computer Name",
                        "sn": "SN",
                        "symptomEnName": "Fault Description"
                    })[["SN", "Fault Description", "Computer Name"]]
 
                    top_computers = detail_df["Computer Name"].value_counts().head(3).to_dict()
                    rows = ""
                    for comp, count in top_computers.items():
                        comp_faults = detail_df[detail_df["Computer Name"] == comp]
                        faults = comp_faults["Fault Description"].value_counts().head(3)
                        fault_lines = "".join([f"{i+1}. {fault} → {qty}<br>" for i, (fault, qty) in enumerate(faults.items())])
                        rows += f"<tr><td>{comp} → {count}</td><td>{fault_lines}</td></tr>"
                    ntf_blocks.append(f"""
                        <h3>{station} — NTF Analysis</h3>
                        <table>
                            <thead><tr><th>Top Computer → Qty</th><th>Top 3 Faults → Qty</th></tr></thead>
                            <tbody>{rows}</tbody>
                        </table>
                    """)
 
                if station in Config.DER_GOALS and der is not None and der > Config.DER_GOALS[station]:
                    failed_stations.append((station, "DER", der, Config.DER_GOALS[station]))
                    detail_df = pd.DataFrame(get_station_der_details(token, project, station)).rename(columns={
                        "sn": "SN",
                        "responsibilityEnName": "Responsibility",
                        "symptomEnName": "Symptoms"
                    })[["SN", "Responsibility", "Symptoms"]]
 
                    top_symptoms = detail_df["Symptoms"].value_counts().head(3)
                    top_responsibilities = detail_df["Responsibility"].value_counts().head(3)
 
                    rows = ""
                    for i in range(3):
                        symptom = top_symptoms.index[i] if i < len(top_symptoms) else ""
                        symptom_qty = top_symptoms.iloc[i] if i < len(top_symptoms) else ""
                        resp = top_responsibilities.index[i] if i < len(top_responsibilities) else ""
                        resp_qty = top_responsibilities.iloc[i] if i < len(top_responsibilities) else ""
                        rows += f"<tr><td>{symptom} → {symptom_qty}</td><td>{resp} → {resp_qty}</td></tr>"
                    der_blocks.append(f"""
                        <h3>{station} — DER Analysis</h3>
                        <table>
                            <thead><tr><th>Symptom → Qty</th><th>Responsibility → Qty</th></tr></thead>
                            <tbody>{rows}</tbody>
                        </table>
                    """)
    except Exception as e:
        print("RTY analysis error:", e)
 
    # Build FPY table manually with wider project column
    fpy_html = "<table border='1' cellspacing='0' cellpadding='4' style='width:100%;'>"
    fpy_html += "<thead><tr>"
    for col in fpy_df.columns:
        if col == "project":
            fpy_html += f"<th style='font-size:10px; white-space:normal; word-wrap:break-word; width:120px;'>{col}</th>"
        else:
            fpy_html += f"<th style='font-size:10px; white-space:normal; word-wrap:break-word;'>{col}</th>"
    fpy_html += "</tr></thead><tbody>"
 
    for _, row in fpy_df.iterrows():
        fpy_html += "<tr>"
        for col in fpy_df.columns:
            if col == "project":
                fpy_html += f"<td style='font-size:10px; white-space:normal; word-wrap:break-word; width:120px;'>{row[col]}</td>"
            else:
                fpy_html += f"<td style='font-size:10px; white-space:normal; word-wrap:break-word;'>{row[col]}</td>"
        fpy_html += "</tr>"
    fpy_html += "</tbody></table>"
 
    # Build full HTML
    html = f"""
    <html><head><style>
    body {{ font-family: Arial; font-size: 11px; }}
    table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; }}
    th, td {{ border: 1px solid #ccc; padding: 4px; text-align: left; vertical-align: top; }}
    th {{ background-color: #4361ee; color: white; }}
    h1 {{ margin-bottom: 10px; color: #4361ee; }}
    h2 {{ margin-top: 30px; color: #4361ee; }}
    h3 {{ margin-top: 20px; color: #4361ee; }}
    </style></head><body>
    <h1>FPY Report for {project}</h1>
    <p><strong>RTY Goal:</strong> {rty_goal}%</p>
 
    <h2>FPY Table</h2>
    {fpy_html}
    """
 
    if failed_stations:
        fail_df = pd.DataFrame(failed_stations, columns=["Station", "Metric", "Actual (%)", "Goal (%)"])
        html += "<h2>Failed Stations</h2>" + fail_df.to_html(index=False)
 
    if ntf_blocks or der_blocks:
        html += "<h2>Top Failure Analysis</h2>"
        html += "".join(ntf_blocks)
        html += "".join(der_blocks)
 
    html += "</body></html>"
 
    # Convert to PDF using ReportLab
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch
 
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    heading_style = styles['Heading1']
    normal_style = styles['Normal']
    
    # Title
    elements.append(Paragraph(f"FPY Report for {project}", title_style))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"RTY Goal: {rty_goal}%", normal_style))
    elements.append(Spacer(1, 12))
    
    # FPY Table
    elements.append(Paragraph("FPY Table", heading_style))
    elements.append(Spacer(1, 12))
    
    # Convert DataFrame to list of lists for ReportLab
    fpy_data_list = [list(fpy_df.columns)] + fpy_df.values.tolist()
    fpy_table = Table(fpy_data_list)
    fpy_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(fpy_table)
    elements.append(Spacer(1, 12))
    
    # Failed Stations
    if failed_stations:
        elements.append(Paragraph("Failed Stations", heading_style))
        elements.append(Spacer(1, 12))
        
        failed_data = [["Station", "Metric", "Actual (%)", "Goal (%)"]] + failed_stations
        failed_table = Table(failed_data)
        failed_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(failed_table)
        elements.append(Spacer(1, 12))
    
    # NTF Breakdown
    if ntf_blocks:
        elements.append(Paragraph("Top Failure Analysis — NTF", heading_style))
        elements.append(Spacer(1, 12))
        
        for rows in ntf_blocks:
            ntf_table = Table([["Top Computer → Qty", "Top 3 Faults → Qty"]] + rows)
            ntf_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(ntf_table)
            elements.append(Spacer(1, 12))
    
    # DER Breakdown
    if der_blocks:
        elements.append(Paragraph("Top Failure Analysis — DER", heading_style))
        elements.append(Spacer(1, 12))
        
        for rows in der_blocks:
            der_table = Table([["Symptom → Qty", "Responsibility → Qty"]] + rows)
            der_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(der_table)
            elements.append(Spacer(1, 12))
    
    # Build PDF
    doc.build(elements)
    
    buffer.seek(0)
    return send_file(buffer,
                     download_name=f"{project}_full_report.pdf",
                     as_attachment=True,
                     mimetype='application/pdf')

@dashboard_bp.route('/export-excel-model')
@login_required
def export_excel_model():
    model_name = request.args.get('model_name')
    station_type = request.args.get('station_type', 'BE')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    rty_goal = float(request.args.get('rty_goal', 90.0))

    try:
        # Try to get data from database first
        fpy_data_query = FPYData.query.filter_by(project=model_name)
        
        # If date range is specified, filter by date
        if start_date and end_date:
            start_datetime = datetime.strptime(start_date, '%Y-%m-%dT%H:%M')
            end_datetime = datetime.strptime(end_date, '%Y-%m-%dT%H:%M')
            fpy_data_query = fpy_data_query.filter(FPYData.timestamp.between(start_datetime, end_datetime))
        
        db_data = fpy_data_query.all()
        
        # If no data in database, fetch from API
        if not db_data:
            token = get_token()
            fpy_data_raw = get_fpy_by_model(token, model_name, station_type, start_date, end_date)
            
            # Convert API data to database format and save
            for record in fpy_data_raw:
                mapped_data = map_api_to_db(record)
                new_record = FPYData(**mapped_data)
                db.session.add(new_record)
            
            db.session.commit()
            
            # Now get the data from the database
            fpy_data_query = FPYData.query.filter_by(project=model_name)
            if start_date and end_date:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%dT%H:%M')
                end_datetime = datetime.strptime(end_date, '%Y-%m-%dT%H:%M')
                fpy_data_query = fpy_data_query.filter(FPYData.timestamp.between(start_datetime, end_datetime))
            
            db_data = fpy_data_query.all()
        
        # Convert database objects to dictionaries
        desired_columns = ["project", "station", "inPut", "pass", "fail", "notFail", "der", "ntf", "rty", "py"]
        fpy_data = []
        
        for record in db_data:
            row = map_db_to_api(record)
            fpy_data.append(row)

        if not fpy_data:
            return "No data to export."

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "FPY Report"

        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Write headers
        headers = ["Model", "Station", "Input Qty", "Good Qty", "NG", "NDF", "NG Rate", "NDF Rate", "RTY", "PY"]
        ws.append(headers)
        
        for cell in ws[1]:  # First row
            cell.font = header_font
            cell.alignment = center
            cell.fill = header_fill

        # Write data
        for row_data in fpy_data:
            row = [
                row_data.get("project", ""),
                row_data.get("station", ""),
                row_data.get("inPut", ""),
                row_data.get("pass", ""),
                row_data.get("fail", ""),
                row_data.get("notFail", ""),
                row_data.get("der", ""),
                row_data.get("ntf", ""),
                row_data.get("rty", ""),
                row_data.get("py", "")
            ]
            ws.append(row)

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output,
                         download_name=f"{model_name}_report.xlsx",
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"Error exporting Excel: {e}")
        return f"Error exporting data: {str(e)}"

@dashboard_bp.route('/export-pdf-model')
@login_required
def export_pdf_model():
    model_name = request.args.get('model_name')
    station_type = request.args.get('station_type', 'BE')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    rty_goal = float(request.args.get('rty_goal', 90.0))

    try:
        # Try to get data from database first
        fpy_data_query = FPYData.query.filter_by(project=model_name)
        
        # If date range is specified, filter by date
        if start_date and end_date:
            start_datetime = datetime.strptime(start_date, '%Y-%m-%dT%H:%M')
            end_datetime = datetime.strptime(end_date, '%Y-%m-%dT%H:%M')
            fpy_data_query = fpy_data_query.filter(FPYData.timestamp.between(start_datetime, end_datetime))
        
        db_data = fpy_data_query.all()
        
        # If no data in database, fetch from API
        if not db_data:
            token = get_token()
            fpy_data_raw = get_fpy_by_model(token, model_name, station_type, start_date, end_date)
            
            # Convert API data to database format and save
            for record in fpy_data_raw:
                mapped_data = map_api_to_db(record)
                new_record = FPYData(**mapped_data)
                db.session.add(new_record)
            
            db.session.commit()
            
            # Now get the data from the database
            fpy_data_query = FPYData.query.filter_by(project=model_name)
            if start_date and end_date:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%dT%H:%M')
                end_datetime = datetime.strptime(end_date, '%Y-%m-%dT%H:%M')
                fpy_data_query = fpy_data_query.filter(FPYData.timestamp.between(start_datetime, end_datetime))
            
            db_data = fpy_data_query.all()
        
        # Convert database objects to dictionaries
        desired_columns = ["project", "station", "inPut", "pass", "fail", "notFail", "der", "ntf", "rty", "py"]
        fpy_data = []
        
        for record in db_data:
            row = map_db_to_api(record)
            fpy_data.append(row)

        if not fpy_data:
            return "No data to export."

        # Create PDF using ReportLab
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.units import inch

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        heading_style = styles['Heading1']
        normal_style = styles['Normal']
        
        # Title
        elements.append(Paragraph(f"FPY Report for {model_name}", title_style))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"RTY Goal: {rty_goal}%", normal_style))
        elements.append(Spacer(1, 12))
        
        # FPY Table
        elements.append(Paragraph("FPY Data", heading_style))
        elements.append(Spacer(1, 12))
        
        # Convert DataFrame to list of lists for ReportLab
        headers = ["Model", "Station", "Input Qty", "Good Qty", "NG", "NDF", "NG Rate", "NDF Rate", "RTY", "PY"]
        fpy_data_list = [headers]
        
        for row in fpy_data:
            fpy_data_list.append([
                row.get("project", ""),
                row.get("station", ""),
                row.get("inPut", ""),
                row.get("pass", ""),
                row.get("fail", ""),
                row.get("notFail", ""),
                row.get("der", ""),
                row.get("ntf", ""),
                row.get("rty", ""),
                row.get("py", "")
            ])
        
        fpy_table = Table(fpy_data_list)
        fpy_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(fpy_table)
        elements.append(Spacer(1, 12))
        
        # Build PDF
        doc.build(elements)
        
        buffer.seek(0)
        return send_file(buffer,
                         download_name=f"{model_name}_report.pdf",
                         as_attachment=True,
                         mimetype='application/pdf')
    except Exception as e:
        print(f"Error exporting PDF: {e}")
        return f"Error exporting data: {str(e)}"
    
    
    
@dashboard_bp.route('/models-rty-summary')
@login_required
def models_rty_summary():
    current_time = datetime.now().strftime('%H:%M')
    try:
        # Import station goals from config
        from config import Config
        
        # Get token
        token = get_token()
        
        # Get project list
        projects = get_project_list(token)
        
        # Get today's data from 08:00AM to now
        now = datetime.now()
        start = now.replace(hour=8, minute=0, second=0, microsecond=0)
        
        # Get FPY data
        fpy_data_raw = get_fpy(token, projects, start, now)
        
        # Get model descriptions from database to filter out 2G models and get goals
        db_models = {model.model_name: model for model in ModelDescription.query.all()}
        
        # Process data and filter out 2G models
        all_models_data = []
        models_achieving_goal = []
        models_not_achieving_goal = []
        
        # Group data by project/model
        project_data = {}
        station_data = {}  # For station analysis
        
        # Initialize counters for overall RTY calculation
        total_pcurr_input_all = 0
        total_ng_all = 0
        total_ndf_all = 0
        
        # Initialize data structures for technology and position analysis
        tech_rty_data = {}  # To store RTY values for each technology
        position_rty_data = {}  # To store RTY values for each position category
        
        for record in fpy_data_raw:
            project = record.get('project', '')
            station = record.get('station', '')
            model_info = db_models.get(project)
            
            # Skip 2G models
            if model_info and model_info.technology == '2G':
                continue
                
            # Skip if no model info found
            if not model_info:
                continue
                
            # Extract values
            input_qty = int(record.get('inPut', 0))
            pass_qty = int(record.get('pass', 0))
            fail_qty = int(record.get('fail', 0))
            not_fail_qty = int(record.get('notFail', 0))
            der_val = float(str(record.get('der', '0')).replace('%', '')) if record.get('der') else 0
            ntf_val = float(str(record.get('ntf', '0')).replace('%', '')) if record.get('ntf') else 0
            rty_val = float(str(record.get('rty', '0')).replace('%', '')) if record.get('rty') else 0
            
            # Get goal from model info
            goal_str = model_info.goal.replace('%', '') if model_info.goal and model_info.goal != 'NA' else '0'
            goal_val = float(goal_str) if goal_str else 0
            
            # Create processed record
            processed_record = {
                "project": project,
                "station": station,
                "inPut": input_qty,
                "pass": pass_qty,
                "fail": fail_qty,
                "notFail": not_fail_qty,
                "der": der_val,
                "ntf": ntf_val,
                "rty": rty_val,
                "goal": goal_val,
                "technology": model_info.technology,
                "position": model_info.position,
                "brand": model_info.brand
            }
            
            # Add to all models data
            all_models_data.append(processed_record)
            
            # Group by project for later analysis
            if project not in project_data:
                project_data[project] = []
            project_data[project].append(processed_record)
            
            # Group by station for station analysis
            if station not in station_data:
                station_data[station] = []
            station_data[station].append(processed_record)
            
            # Accumulate for overall RTY calculation
            total_ng_all += fail_qty
            total_ndf_all += not_fail_qty
            if station == 'PCURR':
                total_pcurr_input_all += input_qty
            
            # Collect RTY data for technology and position analysis
            tech = model_info.technology
            if tech not in tech_rty_data:
                tech_rty_data[tech] = []
            tech_rty_data[tech].append(rty_val)
            
            position = model_info.position
            if position in ['LOW', 'MID']:
                position_key = 'LOW/MID'
            elif position in ['HIGH', 'FLAGSHIP', 'FOLD']:
                position_key = 'HIGH/FLAGSHIP/FOLD'
            else:
                position_key = 'OTHER'
                
            if position_key not in position_rty_data:
                position_rty_data[position_key] = []
            position_rty_data[position_key].append(rty_val)
        
        # Calculate overall RTY using the formula
        # RTY (%) = (1 - (Total NG Qty + Total NDF Qty) / Total PCURR Input) × 100
        overall_rty = (1 - (total_ng_all + total_ndf_all) / total_pcurr_input_all) * 100 if total_pcurr_input_all > 0 else 0
        
        # Calculate average RTY by technology and position
        tech_avg_rty = {}
        for tech, rty_values in tech_rty_data.items():
            if rty_values:
                tech_avg_rty[tech] = sum(rty_values) / len(rty_values)
            else:
                tech_avg_rty[tech] = 0
        
        position_avg_rty = {}
        for position, rty_values in position_rty_data.items():
            if rty_values:
                position_avg_rty[position] = sum(rty_values) / len(rty_values)
            else:
                position_avg_rty[position] = 0
        
        # Calculate RTY for each project and check against goal
        for project, records in project_data.items():
            # Calculate total input and pass for the project
            total_input = sum(record['inPut'] for record in records)
            total_pass = sum(record['pass'] for record in records)
            
            # Calculate overall RTY for the project (from API)
            # We'll use the RTY from the first record as the overall RTY for the project
            project_rty = records[0]['rty'] if records else 0
            
            # Get model info
            model_info = db_models.get(project)
            goal_val = model_info.goal.replace('%', '') if model_info.goal and model_info.goal != 'NA' else '0'
            goal_val = float(goal_val) if goal_val else 0
            
            # Check if goal is achieved
            if project_rty >= goal_val and goal_val > 0:
                models_achieving_goal.append({
                    "project": project,
                    "technology": model_info.technology,
                    "position": model_info.position,
                    "brand": model_info.brand,
                    "goal": goal_val,
                    "actual_rty": project_rty,
                    "gap": project_rty - goal_val,
                    "total_input": total_input,
                    "total_pass": total_pass,
                    "total_fail": sum(record['fail'] for record in records),
                    "total_ntf": sum(record['notFail'] for record in records),
                    "stations": records
                })
            else:
                models_not_achieving_goal.append({
                    "project": project,
                    "technology": model_info.technology,
                    "position": model_info.position,
                    "brand": model_info.brand,
                    "goal": goal_val,
                    "actual_rty": project_rty,
                    "gap": goal_val - project_rty if goal_val > 0 else 0,
                    "total_input": total_input,
                    "total_pass": total_pass,
                    "total_fail": sum(record['fail'] for record in records),
                    "total_ntf": sum(record['notFail'] for record in records),
                    "stations": records
                })
        
        # Sort models not achieving goal by gap (largest gap first)
        models_not_achieving_goal.sort(key=lambda x: x['gap'], reverse=True)
        
        # Calculate station performance for models not achieving goal
        stations_not_meeting_goal = {}
        station_details = {}  # For storing top 3 issues
        
        for model in models_not_achieving_goal:
            project = model['project']
            stations_not_meeting_goal[project] = []
            station_details[project] = []
            
            # Get stations for this model
            for station, records in station_data.items():
                # Filter records for this project
                project_records = [r for r in records if r['project'] == project]
                
                if project_records:
                    total_input = sum(record['inPut'] for record in project_records)
                    total_pass = sum(record['pass'] for record in project_records)
                    total_fail = sum(record['fail'] for record in project_records)
                    total_ntf = sum(record['notFail'] for record in project_records)
                    
                    # Calculate station RTY
                    station_rty = (total_pass / total_input * 100) if total_input > 0 else 0
                    
                    # Calculate station DER and NTF
                    station_der = (total_fail / total_input * 100) if total_input > 0 else 0
                    station_ntf = (total_ntf / total_input * 100) if total_input > 0 else 0
                    
                    # Check if station meets goals from config
                    station_ntf_goal = Config.NTF_GOALS.get(station, float('inf'))
                    station_der_goal = Config.DER_GOALS.get(station, float('inf'))
                    
                    # Consider a station not meeting goal if it exceeds either NTF or DER goal
                    if station_ntf > station_ntf_goal or station_der > station_der_goal:
                        station_info = {
                            "station": station,
                            "rty": station_rty,
                            "der": station_der,
                            "ntf": station_ntf,
                            "der_goal": station_der_goal,
                            "ntf_goal": station_ntf_goal,
                            "der_gap": station_der - station_der_goal,
                            "ntf_gap": station_ntf - station_ntf_goal,
                            "total_input": total_input,
                            "total_pass": total_pass,
                            "total_fail": total_fail,
                            "total_ntf": total_ntf
                        }
                        
                        stations_not_meeting_goal[project].append(station_info)
                        
                        # Get NTF details if NTF goal is not met
                        if station_ntf > station_ntf_goal:
                            try:
                                # Get detailed NTF data for this station
                                ntf_data = get_station_ntf_details(token, project, station)
                                ntf_df = pd.DataFrame(ntf_data)
                                
                                if not ntf_df.empty:
                                    # Rename columns to match expected format
                                    ntf_df = ntf_df.rename(columns={
                                        "substation": "Computer Name",
                                        "sn": "SN",
                                        "symptomEnName": "Fault Description"
                                    })
                                    
                                    # Get top computers
                                    top_computers = ntf_df["Computer Name"].value_counts().head(3).to_dict()
                                    
                                    # Get top faults by computer
                                    top_faults_by_computer = {}
                                    for comp in top_computers:
                                        comp_faults = ntf_df[ntf_df["Computer Name"] == comp]
                                        faults = comp_faults["Fault Description"].value_counts().head(3).reset_index().values.tolist()
                                        top_faults_by_computer[comp] = faults
                                    
                                    station_info["ntf_details"] = {
                                        "top_computers": top_computers,
                                        "top_faults_by_computer": top_faults_by_computer
                                    }
                                else:
                                    station_info["ntf_details"] = {
                                        "top_computers": {},
                                        "top_faults_by_computer": {}
                                    }
                            except Exception as e:
                                print(f"Error getting NTF details for {station}: {e}")
                                station_info["ntf_details"] = {
                                    "top_computers": {},
                                    "top_faults_by_computer": {}
                                }
                        
                        # Get DER details if DER goal is not met
                        if station_der > station_der_goal:
                            try:
                                # Get detailed DER data for this station
                                der_data = get_station_der_details(token, project, station)
                                der_df = pd.DataFrame(der_data)
                                
                                if not der_df.empty:
                                    # Rename columns to match expected format
                                    der_df = der_df.rename(columns={
                                        "sn": "SN",
                                        "responsibilityEnName": "Responsibility",
                                        "symptomEnName": "Symptoms"
                                    })
                                    
                                    # Get top symptoms and responsibilities
                                    top_symptoms = get_top_n_counts(der_df, "Symptoms", 3)
                                    top_responsibilities = get_top_n_counts(der_df, "Responsibility", 3)
                                    
                                    station_info["der_details"] = {
                                        "top_symptoms": top_symptoms.to_dict(orient="records"),
                                        "top_responsibilities": top_responsibilities.to_dict(orient="records")
                                    }
                                else:
                                    station_info["der_details"] = {
                                        "top_symptoms": [],
                                        "top_responsibilities": []
                                    }
                            except Exception as e:
                                print(f"Error getting DER details for {station}: {e}")
                                station_info["der_details"] = {
                                    "top_symptoms": [],
                                    "top_responsibilities": []
                                }
            
            # Sort stations not meeting goal by DER gap (largest gap first)
            stations_not_meeting_goal[project].sort(key=lambda x: x['der_gap'], reverse=True)
        
        # Calculate summary statistics
        total_models = len(project_data)
        models_achieving_count = len(models_achieving_goal)
        models_not_achieving_count = len(models_not_achieving_goal)
        achievement_rate = (models_achieving_count / total_models * 100) if total_models > 0 else 0
        
        # Prepare summary data
        summary_data = {
            "total_models": total_models,
            "models_achieving_goal": models_achieving_count,
            "models_not_achieving_goal": models_not_achieving_count,
            "achievement_rate": achievement_rate,
            "overall_rty": overall_rty,
            "total_pcurr_input": total_pcurr_input_all,
            "total_ng_all": total_ng_all,
            "total_ndf_all": total_ndf_all,
            "ntf_goals": Config.NTF_GOALS,
            "der_goals": Config.DER_GOALS,
            "tech_avg_rty": tech_avg_rty,
            "position_avg_rty": position_avg_rty
        }
        
        # Prepare models summary table data
        models_summary = []
        
        # Add models achieving goal
        for model in models_achieving_goal:
            models_summary.append({
                "project": model["project"],
                "technology": model["technology"],
                "position": model["position"],
                "brand": model["brand"],
                "goal": model["goal"],
                "actual_rty": model["actual_rty"],
                "status": "Achieved",
                "gap": model["gap"]
            })
        
        # Add models not achieving goal
        for model in models_not_achieving_goal:
            models_summary.append({
                "project": model["project"],
                "technology": model["technology"],
                "position": model["position"],
                "brand": model["brand"],
                "goal": model["goal"],
                "actual_rty": model["actual_rty"],
                "status": "Not Achieved",
                "gap": -model["gap"]  # Negative gap for not achieved
            })
        
        # Sort by project name
        models_summary.sort(key=lambda x: x["project"])
        
        return render_template('dashboard/models_rty_summary.html', 
                               models_summary=models_summary,
                               models_achieving_goal=models_achieving_goal,
                               models_not_achieving_goal=models_not_achieving_goal,
                               stations_not_meeting_goal=stations_not_meeting_goal,
                               station_details=station_details,
                               summary_data=summary_data,
                               current_time=current_time)
    except Exception as e:
        current_time = datetime.now().strftime('%H:%M')
        return render_template('errors/500.html', error=str(e), current_time=current_time)
        
        
        
        
@dashboard_bp.route('/model-analysis-date-range', methods=['GET', 'POST'])
@login_required
def model_analysis_date_range():
    current_time = datetime.now().strftime('%H:%M')
    try:
        # Get all models from database
        models = ModelDescription.query.all()
        
        # Initialize variables
        selected_model = None
        station_type = "BE"
        start_date = None
        end_date = None
        rty_goal = None
        auto_goal = None
        data = []
        failed_stations = []
        fail_details = []
        date_range = "Last Day"  # Default selection
        
        if request.method == 'POST':
            selected_model = request.form.get('model_name')
            station_type = request.form.get('station_type', 'BE')
            rty_goal = float(request.form.get('rty_goal', 0))
            auto_goal = request.form.get('auto_goal')
            date_range = request.form.get('date_range', 'Last Day')
            
            # Calculate date range based on selection
            now = datetime.now()
            
            if date_range == "Last Day":
                # Yesterday 08:00 AM to today 08:00 AM
                end_date = now.replace(hour=8, minute=0, second=0, microsecond=0)
                start_date = end_date - timedelta(days=1)
            elif date_range == "Last Week":
                # Previous week's Monday 08:00 AM to this week's Monday 08:00 AM
                today_weekday = now.weekday()  # Monday is 0, Sunday is 6
                days_since_monday = today_weekday
                this_monday = now.replace(hour=8, minute=0, second=0, microsecond=0) - timedelta(days=days_since_monday)
                start_date = this_monday - timedelta(weeks=1)
                end_date = this_monday
            elif date_range == "This Week":
                # This week's Monday 08:00 AM to next Monday 08:00 AM (or now if next Monday hasn't come)
                today_weekday = now.weekday()
                days_since_monday = today_weekday
                this_monday = now.replace(hour=8, minute=0, second=0, microsecond=0) - timedelta(days=days_since_monday)
                start_date = this_monday
                end_date = min(this_monday + timedelta(weeks=1), now)
            elif date_range == "Last Month":
                # Previous month's first day 08:00 AM to this month's first day 08:00 AM
                first_day_current_month = now.replace(day=1, hour=8, minute=0, second=0, microsecond=0)
                last_day_prev_month = first_day_current_month - timedelta(days=1)
                start_date = last_day_prev_month.replace(day=1, hour=8, minute=0, second=0, microsecond=0)
                end_date = first_day_current_month
            elif date_range == "This Month":
                # This month's first day 08:00 AM to next month's first day 08:00 AM (or now if next month hasn't come)
                first_day_current_month = now.replace(day=1, hour=8, minute=0, second=0, microsecond=0)
                if now.month == 12:
                    next_month = now.replace(year=now.year+1, month=1, day=1)
                else:
                    next_month = now.replace(month=now.month+1, day=1)
                start_date = first_day_current_month
                end_date = min(next_month.replace(hour=8, minute=0, second=0, microsecond=0), now)
            
            # Format dates for display
            start_date_str = start_date.strftime('%Y-%m-%dT%H:%M')
            end_date_str = end_date.strftime('%Y-%m-%dT%H:%M')
            
            # Get token
            token = get_token()
            
            # Get FPY data for the selected model and date range
            fpy_data = get_fpy_by_model(
                token, 
                selected_model, 
                station_type, 
                start_date_str, 
                end_date_str
            )
            
            # Process data
            if fpy_data:
                # Get model info
                model_info = ModelDescription.query.filter_by(model_name=selected_model).first()
                
                # Calculate overall metrics
                total_input = sum(int(record.get('inPut', 0)) for record in fpy_data)
                total_pass = sum(int(record.get('pass', 0)) for record in fpy_data)
                total_fail = sum(int(record.get('fail', 0)) for record in fpy_data)
                total_ntf = sum(int(record.get('notFail', 0)) for record in fpy_data)
                
                # Calculate RTY
                overall_rty = (total_pass / total_input * 100) if total_input > 0 else 0
                
                # Check if goal is achieved
                goal_achieved = overall_rty >= rty_goal if rty_goal > 0 else True
                
                # Prepare data for display
                data = []
                for record in fpy_data:
                    data.append({
                        "project": record.get('project'),
                        "station": record.get('station'),
                        "inPut": record.get('inPut'),
                        "pass": record.get('pass'),
                        "fail": record.get('fail'),
                        "notFail": record.get('notFail'),
                        "der": record.get('der'),
                        "ntf": record.get('ntf'),
                        "rty": record.get('rty'),
                        "py": record.get('py')
                    })
                
                # Check for stations not meeting goals
                from config import Config
                failed_stations = []
                fail_details = []
                
                for record in fpy_data:
                    station = record.get('station')
                    der_val = float(str(record.get('der', '0')).replace('%', '')) if record.get('der') else 0
                    ntf_val = float(str(record.get('ntf', '0')).replace('%', '')) if record.get('ntf') else 0
                    
                    # Check if station meets goals
                    station_ntf_goal = Config.NTF_GOALS.get(station, float('inf'))
                    station_der_goal = Config.DER_GOALS.get(station, float('inf'))
                    
                    if ntf_val > station_ntf_goal:
                        failed_stations.append((station, "NTF", f"{ntf_val:.2f}%", f"{station_ntf_goal:.2f}%"))
                        
                        # Get NTF details
                        try:
                            ntf_details = get_station_ntf_details_by_model(
                                token, selected_model, station, station_type, 
                                start_date_str, end_date_str
                            )
                            
                            if ntf_details:
                                ntf_df = pd.DataFrame(ntf_details)
                                if not ntf_df.empty:
                                    ntf_df = ntf_df.rename(columns={
                                        "substation": "Computer Name",
                                        "sn": "SN",
                                        "symptomEnName": "Fault Description"
                                    })
                                    
                                    top_computers = ntf_df["Computer Name"].value_counts().head(3).to_dict()
                                    top_faults_by_computer = {}
                                    
                                    for comp in top_computers:
                                        comp_faults = ntf_df[ntf_df["Computer Name"] == comp]
                                        faults = comp_faults["Fault Description"].value_counts().head(3).reset_index().values.tolist()
                                        top_faults_by_computer[comp] = faults
                                    
                                    fail_details.append({
                                        "station": station,
                                        "metric": "NTF",
                                        "actual": ntf_val,
                                        "goal": station_ntf_goal,
                                        "top_computers": top_computers,
                                        "top_faults_by_computer": top_faults_by_computer
                                    })
                        except Exception as e:
                            print(f"Error getting NTF details: {e}")
                    
                    if der_val > station_der_goal:
                        failed_stations.append((station, "DER", f"{der_val:.2f}%", f"{station_der_goal:.2f}%"))
                        
                        # Get DER details
                        try:
                            der_details = get_station_der_details_by_model(
                                token, selected_model, station, station_type, 
                                start_date_str, end_date_str
                            )
                            
                            if der_details:
                                der_df = pd.DataFrame(der_details)
                                if not der_df.empty:
                                    der_df = der_df.rename(columns={
                                        "sn": "SN",
                                        "responsibilityEnName": "Responsibility",
                                        "symptomEnName": "Symptoms"
                                    })
                                    
                                    top_symptoms = get_top_n_counts(der_df, "Symptoms", 3)
                                    top_responsibilities = get_top_n_counts(der_df, "Responsibility", 3)
                                    
                                    fail_details.append({
                                        "station": station,
                                        "metric": "DER",
                                        "actual": der_val,
                                        "goal": station_der_goal,
                                        "top_symptoms": top_symptoms.to_dict(orient="records"),
                                        "top_responsibilities": top_responsibilities.to_dict(orient="records")
                                    })
                        except Exception as e:
                            print(f"Error getting DER details: {e}")
        
        return render_template(
            'dashboard/model_analysis_date_range.html',
            models=models,
            selected_model=selected_model,
            station_type=station_type,
            start_date=start_date_str if start_date else '',
            end_date=end_date_str if end_date else '',
            rty_goal=rty_goal,
            auto_goal=auto_goal,
            data=data,
            failed_stations=failed_stations,
            fail_details=fail_details,
            date_range=date_range,
            current_time=current_time
        )
    except Exception as e:
        current_time = datetime.now().strftime('%H:%M')
        return render_template('errors/500.html', error=str(e), current_time=current_time)
        
        
        
        
@dashboard_bp.route('/export-excel-model-date-range')
@login_required
def export_excel_model_date_range():
    try:
        model_name = request.args.get('model_name')
        station_type = request.args.get('station_type', 'BE')
        date_range = request.args.get('date_range', 'Last Day')
        rty_goal = float(request.args.get('rty_goal', 0))
        
        # Calculate date range based on selection
        now = datetime.now()
        
        if date_range == "Last Day":
            end_date = now.replace(hour=8, minute=0, second=0, microsecond=0)
            start_date = end_date - timedelta(days=1)
        elif date_range == "Last Week":
            today_weekday = now.weekday()
            days_since_monday = today_weekday
            this_monday = now.replace(hour=8, minute=0, second=0, microsecond=0) - timedelta(days=days_since_monday)
            start_date = this_monday - timedelta(weeks=1)
            end_date = this_monday
        elif date_range == "This Week":
            today_weekday = now.weekday()
            days_since_monday = today_weekday
            this_monday = now.replace(hour=8, minute=0, second=0, microsecond=0) - timedelta(days=days_since_monday)
            start_date = this_monday
            end_date = min(this_monday + timedelta(weeks=1), now)
        elif date_range == "Last Month":
            first_day_current_month = now.replace(day=1, hour=8, minute=0, second=0, microsecond=0)
            last_day_prev_month = first_day_current_month - timedelta(days=1)
            start_date = last_day_prev_month.replace(day=1, hour=8, minute=0, second=0, microsecond=0)
            end_date = first_day_current_month
        elif date_range == "This Month":
            first_day_current_month = now.replace(day=1, hour=8, minute=0, second=0, microsecond=0)
            if now.month == 12:
                next_month = now.replace(year=now.year+1, month=1, day=1)
            else:
                next_month = now.replace(month=now.month+1, day=1)
            start_date = first_day_current_month
            end_date = min(next_month.replace(hour=8, minute=0, second=0, microsecond=0), now)
        
        # Format dates
        start_date_str = start_date.strftime('%Y-%m-%dT%H:%M')
        end_date_str = end_date.strftime('%Y-%m-%dT%H:%M')
        
        # Get token and data
        token = get_token()
        fpy_data = get_fpy_by_model(token, model_name, station_type, start_date_str, end_date_str)
        
        # Create Excel file
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Model Analysis"
        
        # Write headers
        headers = ["Station", "Input Qty", "Good Qty", "NG", "NDF", "NG Rate", "NDF Rate", "RTY"]
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)
            worksheet.cell(row=1, column=col_num).font = Font(bold=True)
        
        # Write data
        for row_num, record in enumerate(fpy_data, 2):
            worksheet.cell(row=row_num, column=1, value=record.get('station'))
            worksheet.cell(row=row_num, column=2, value=record.get('inPut'))
            worksheet.cell(row=row_num, column=3, value=record.get('pass'))
            worksheet.cell(row=row_num, column=4, value=record.get('fail'))
            worksheet.cell(row=row_num, column=5, value=record.get('notFail'))
            worksheet.cell(row=row_num, column=6, value=record.get('der'))
            worksheet.cell(row=row_num, column=7, value=record.get('ntf'))
            worksheet.cell(row=row_num, column=8, value=record.get('rty'))
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{model_name}_analysis_{date_range.replace(' ', '_')}.xlsx"
        )
    except Exception as e:
        return f"Error exporting data: {str(e)}", 500
 
@dashboard_bp.route('/export-pdf-model-date-range')
@login_required
def export_pdf_model_date_range():
    try:
        model_name = request.args.get('model_name')
        station_type = request.args.get('station_type', 'BE')
        date_range = request.args.get('date_range', 'Last Day')
        rty_goal = float(request.args.get('rty_goal', 0))
        
        # Calculate date range based on selection
        now = datetime.now()
        
        if date_range == "Last Day":
            end_date = now.replace(hour=8, minute=0, second=0, microsecond=0)
            start_date = end_date - timedelta(days=1)
        elif date_range == "Last Week":
            today_weekday = now.weekday()
            days_since_monday = today_weekday
            this_monday = now.replace(hour=8, minute=0, second=0, microsecond=0) - timedelta(days=days_since_monday)
            start_date = this_monday - timedelta(weeks=1)
            end_date = this_monday
        elif date_range == "This Week":
            today_weekday = now.weekday()
            days_since_monday = today_weekday
            this_monday = now.replace(hour=8, minute=0, second=0, microsecond=0) - timedelta(days=days_since_monday)
            start_date = this_monday
            end_date = min(this_monday + timedelta(weeks=1), now)
        elif date_range == "Last Month":
            first_day_current_month = now.replace(day=1, hour=8, minute=0, second=0, microsecond=0)
            last_day_prev_month = first_day_current_month - timedelta(days=1)
            start_date = last_day_prev_month.replace(day=1, hour=8, minute=0, second=0, microsecond=0)
            end_date = first_day_current_month
        elif date_range == "This Month":
            first_day_current_month = now.replace(day=1, hour=8, minute=0, second=0, microsecond=0)
            if now.month == 12:
                next_month = now.replace(year=now.year+1, month=1, day=1)
            else:
                next_month = now.replace(month=now.month+1, day=1)
            start_date = first_day_current_month
            end_date = min(next_month.replace(hour=8, minute=0, second=0, microsecond=0), now)
        
        # Format dates
        start_date_str = start_date.strftime('%Y-%m-%dT%H:%M')
        end_date_str = end_date.strftime('%Y-%m-%dT%H:%M')
        
        # Get token and data
        token = get_token()
        fpy_data = get_fpy_by_model(token, model_name, station_type, start_date_str, end_date_str)
        
        # Create PDF content (simplified for this example)
        pdf_content = f"""
        Model Analysis Report
        =====================
        
        Model: {model_name}
        Date Range: {date_range}
        Station Type: {station_type}
        RTY Goal: {rty_goal}%
        
        Station Performance Data:
        """
        
        for record in fpy_data:
            pdf_content += f"""
            Station: {record.get('station')}
            Input: {record.get('inPut')}
            Pass: {record.get('pass')}
            Fail: {record.get('fail')}
            NDF: {record.get('notFail')}
            DER: {record.get('der')}
            NTF: {record.get('ntf')}
            RTY: {record.get('rty')}
            
            """
        
        # Create a simple text file (in a real implementation, you would use a PDF library like ReportLab)
        output = io.BytesIO()
        output.write(pdf_content.encode('utf-8'))
        output.seek(0)
        
        return send_file(
            output,
            mimetype='text/plain',
            as_attachment=True,
            download_name=f"{model_name}_analysis_{date_range.replace(' ', '_')}.txt"
        )
    except Exception as e:
        return f"Error exporting PDF: {str(e)}", 500
